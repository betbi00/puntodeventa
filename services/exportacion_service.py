"""Exportación de los reportes a Excel (.xlsx): una hoja por cada reporte,
para el mismo rango de fechas que se ve en pantalla."""
from openpyxl import Workbook
from openpyxl.styles import Font

from services import reporte_service as rep

TIPOS_EXTRAS_BEBIDA = ["boba", "perla_explosiva"]


def exportar_excel(desde: str, hasta: str, ruta_destino: str) -> None:
    """desde/hasta en formato 'YYYY-MM-DD'. Genera el archivo en ruta_destino."""
    wb = Workbook()

    _hoja_resumen(wb.active, desde, hasta)
    _hoja_ventas_por_dia(wb.create_sheet("Ventas por día"), desde, hasta)
    _hoja_productos(wb.create_sheet("Productos más vendidos"), desde, hasta)
    _hoja_empleados(wb.create_sheet("Ventas por empleado"), desde, hasta)
    _hoja_consumo_extras(wb.create_sheet("Consumo boba y perlas"), desde, hasta)

    wb.save(ruta_destino)


def _titulo(hoja, texto):
    hoja["A1"] = texto
    hoja["A1"].font = Font(bold=True, size=14)


def _encabezados(hoja, fila, columnas):
    for col_idx, texto in enumerate(columnas, start=1):
        celda = hoja.cell(row=fila, column=col_idx, value=texto)
        celda.font = Font(bold=True)


def _ajustar_anchos(hoja, anchos):
    for col, ancho in anchos.items():
        hoja.column_dimensions[col].width = ancho


def _hoja_resumen(hoja, desde, hasta):
    hoja.title = "Resumen"
    _titulo(hoja, f"Resumen de ventas · {desde} a {hasta}")
    resumen = rep.resumen_ventas(desde, hasta)
    filas = [
        ("Ventas del período", resumen["num_ventas"]),
        ("Ingresos totales", resumen["ingresos_totales"]),
        ("Ticket promedio", resumen["ticket_promedio"]),
        ("Descuento aplicado", resumen["descuento_total"]),
        ("Ingresos en efectivo", resumen["ingresos_efectivo"]),
        ("Ingresos con tarjeta", resumen["ingresos_tarjeta"]),
    ]
    for i, (etiqueta, valor) in enumerate(filas, start=3):
        hoja.cell(row=i, column=1, value=etiqueta)
        hoja.cell(row=i, column=2, value=valor)
    _ajustar_anchos(hoja, {"A": 24, "B": 16})


def _hoja_ventas_por_dia(hoja, desde, hasta):
    _titulo(hoja, "Ventas por día")
    _encabezados(hoja, 3, ["Día", "Número de ventas", "Ingresos"])
    for i, d in enumerate(rep.ventas_por_dia(desde, hasta), start=4):
        hoja.cell(row=i, column=1, value=d["dia"])
        hoja.cell(row=i, column=2, value=d["num_ventas"])
        hoja.cell(row=i, column=3, value=d["ingresos"])
    _ajustar_anchos(hoja, {"A": 14, "B": 18, "C": 14})


def _hoja_productos(hoja, desde, hasta):
    _titulo(hoja, "Productos más vendidos")
    _encabezados(hoja, 3, ["Producto", "Cantidad vendida", "Ingresos"])
    for i, d in enumerate(rep.productos_mas_vendidos(desde, hasta, limite=1000), start=4):
        hoja.cell(row=i, column=1, value=d["nombre"])
        hoja.cell(row=i, column=2, value=d["cantidad"])
        hoja.cell(row=i, column=3, value=d["ingresos"])
    _ajustar_anchos(hoja, {"A": 26, "B": 16, "C": 14})


def _hoja_empleados(hoja, desde, hasta):
    _titulo(hoja, "Ventas por empleado")
    _encabezados(hoja, 3, ["Empleado", "Número de ventas", "Ingresos"])
    for i, d in enumerate(rep.ventas_por_empleado(desde, hasta), start=4):
        hoja.cell(row=i, column=1, value=d["nombre"])
        hoja.cell(row=i, column=2, value=d["num_ventas"])
        hoja.cell(row=i, column=3, value=d["ingresos"])
    _ajustar_anchos(hoja, {"A": 22, "B": 18, "C": 14})


def _hoja_consumo_extras(hoja, desde, hasta):
    _titulo(hoja, "Consumo de boba y perlas explosivas")
    _encabezados(hoja, 3, ["Insumo", "Cantidad consumida"])
    for i, d in enumerate(rep.consumo_insumos(desde, hasta, tipos=TIPOS_EXTRAS_BEBIDA), start=4):
        hoja.cell(row=i, column=1, value=d["nombre"])
        hoja.cell(row=i, column=2, value=d["cantidad"])
    _ajustar_anchos(hoja, {"A": 26, "B": 18})
